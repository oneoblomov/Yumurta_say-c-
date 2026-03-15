"""
app.py - FastAPI Ana Uygulama
==============================
Tüm route'lar, WebSocket, MJPEG streaming,
HTMX partial render, API endpointleri.
"""

import io
import csv
import json
import asyncio
import subprocess
import os
import socket
from calendar import monthrange
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict
from urllib.parse import urlsplit, urlunsplit

from fastapi import (
    FastAPI, Request, WebSocket, WebSocketDisconnect,
    UploadFile, File, Form, Query, HTTPException,
)
from fastapi.responses import (
    HTMLResponse, JSONResponse, StreamingResponse, FileResponse,
)
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from .database import Database
from .pipeline_manager import PipelineManager
from .update_manager import UpdateManager
from .versioning import display_version, read_version
from .i18n import (
    load_translations, t, get_all_translations,
    SUPPORTED_LANGUAGES, DEFAULT_LANG,
)

# ============================================================ App
BASE_DIR = Path(__file__).resolve().parent
ROOT_DIR = BASE_DIR.parent

app = FastAPI(title="Yumurta Sayıcı", version=read_version())
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")),
          name="static")

templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

# Global instances
db = Database()
pipeline = PipelineManager(db)
update_manager = UpdateManager(db)
schedule_controller_task: Optional[asyncio.Task] = None
schedule_runtime_state = {
    "last_reason": None,
    "last_action": "idle",
    "last_ok": True,
    "last_message": "",
    "last_checked_at": None,
}
_last_schedule_alert_signature = None


def _record_schedule_state(reason: str, action: str,
                           ok: bool, message: str = "") -> Dict[str, object]:
    checked_at = datetime.now().strftime("%H:%M:%S")
    schedule_runtime_state.update({
        "last_reason": reason,
        "last_action": action,
        "last_ok": ok,
        "last_message": message,
        "last_checked_at": checked_at,
    })
    return dict(schedule_runtime_state)


def _maybe_emit_schedule_alert(action: str, message: str) -> None:
    global _last_schedule_alert_signature
    if not message:
        return
    signature = (action, message)
    if signature == _last_schedule_alert_signature:
        return
    _last_schedule_alert_signature = signature
    severity = "warning" if action == "stopped" else "error"
    db.add_alert("camera_schedule", message, severity)


def _status_payload() -> Dict[str, object]:
    status = pipeline.get_status()
    status["schedule_runtime"] = dict(schedule_runtime_state)
    return status


def _enforce_camera_schedule(reason: str = "periodic") -> Dict[str, object]:
    schedule = pipeline.get_schedule_window()
    should_run = pipeline.is_within_schedule()
    status = pipeline.get_status()
    is_running = status.get("running", False)
    is_paused = status.get("paused", False)

    if should_run and not is_running:
        source = db.get_setting("camera_source", "0")
        result = pipeline.start(source=source)
        if result.get("ok"):
            message = (
                f"[SCHEDULE] Pipeline başlatıldı ({reason}) "
                f"{schedule['start']}-{schedule['end']}"
            )
            print(message)
            return _record_schedule_state(reason, "started", True, message)
        else:
            message = (
                f"Pipeline otomatik başlatılamadı ({reason}): "
                f"{result.get('error')}"
            )
            print(f"[SCHEDULE] {message}")
            _maybe_emit_schedule_alert("start_failed", message)
            return _record_schedule_state(reason, "start_failed", False, message)
    elif should_run and is_paused:
        result = pipeline.resume()
        if result.get("ok"):
            message = f"[SCHEDULE] Pipeline devam ettirildi ({reason})"
            print(message)
            return _record_schedule_state(reason, "resumed", True, message)
        message = f"Pipeline otomatik devam ettirilemedi ({reason})"
        _maybe_emit_schedule_alert("resume_failed", message)
        return _record_schedule_state(reason, "resume_failed", False, message)
    elif not should_run and is_running:
        result = pipeline.stop()
        if result.get("ok"):
            message = (
                f"[SCHEDULE] Pipeline durduruldu ({reason}) "
                f"{schedule['start']}-{schedule['end']}"
            )
            print(message)
            return _record_schedule_state(reason, "stopped", True, message)
        else:
            message = (
                f"Pipeline otomatik durdurulamadı ({reason}): "
                f"{result.get('error')}"
            )
            print(f"[SCHEDULE] {message}")
            _maybe_emit_schedule_alert("stop_failed", message)
            return _record_schedule_state(reason, "stop_failed", False, message)

    return _record_schedule_state(reason, "noop", True, "")


async def _camera_schedule_loop() -> None:
    while True:
        try:
            _enforce_camera_schedule(reason="watchdog")
        except Exception as exc:
            print(f"[SCHEDULE] Denetim hatası: {exc}")
            _record_schedule_state("watchdog", "exception", False, str(exc))
        await asyncio.sleep(5)

# Otomatik pipeline denetimi
@app.on_event("startup")
async def startup_event():
    global schedule_controller_task
    print("[STARTUP] Kamera zamanlayıcısı başlatılıyor...")
    if schedule_controller_task is None or schedule_controller_task.done():
        schedule_controller_task = asyncio.create_task(_camera_schedule_loop())
    _enforce_camera_schedule(reason="startup")

    # systemd notification: READY and periodic WATCHDOG pings
    notify_sock = os.environ.get("NOTIFY_SOCKET")
    if notify_sock:
        # send initial READY=1
        try:
            s = socket.socket(socket.AF_UNIX, socket.SOCK_DGRAM)
            if notify_sock[0] == "@":
                notify_sock = "\0" + notify_sock[1:]
            s.connect(notify_sock)
            s.sendall(b"READY=1")
            s.close()
        except Exception:
            pass

        async def _notify_watchdog():
            while True:
                try:
                    s = socket.socket(socket.AF_UNIX, socket.SOCK_DGRAM)
                    if notify_sock[0] == "@":
                        s.connect("\0" + notify_sock[1:])
                    else:
                        s.connect(notify_sock)
                    s.sendall(b"WATCHDOG=1")
                    s.close()
                except Exception:
                    pass
                await asyncio.sleep(30)

        asyncio.create_task(_notify_watchdog())

@app.on_event("shutdown")
async def shutdown_event():
    global schedule_controller_task
    if schedule_controller_task is not None:
        schedule_controller_task.cancel()
        try:
            await schedule_controller_task
        except asyncio.CancelledError:
            pass
        schedule_controller_task = None
    print("[SHUTDOWN] Pipeline durduruluyor...")
    if pipeline.get_status().get("running"):
        pipeline.stop()
    print("[SHUTDOWN] Pipeline durduruldu.")

# WebSocket connections
ws_connections: List[WebSocket] = []


# ============================================================ Helpers

def get_cloudflared_url() -> Optional[str]:
    """Extract the public URL issued by cloudflared from the systemd journal."""
    try:
        output = subprocess.check_output(
            [
                "journalctl", "-u", "cloudflared.service", "-n", "100",
                "--no-pager",
            ],
            text=True,
            stderr=subprocess.DEVNULL,
        )
    except Exception:
        return None

    import re
    pattern = re.compile(r"https?://[\w\-.]+\.trycloudflare\.com[\w\-/]*")
    for line in reversed(output.splitlines()):
        m = pattern.search(line)
        if m:
            parsed = urlsplit(m.group(0))
            return urlunsplit((parsed.scheme, parsed.netloc, "", "", ""))
    return None


def _month_bounds(month_value: Optional[str]) -> tuple[date, date, str, str, str]:
    today = date.today()
    if month_value:
        try:
            month_start = datetime.strptime(month_value, "%Y-%m").date().replace(day=1)
        except ValueError:
            month_start = today.replace(day=1)
    else:
        month_start = today.replace(day=1)

    last_day = monthrange(month_start.year, month_start.month)[1]
    month_end = month_start.replace(day=last_day)

    prev_month = (month_start - timedelta(days=1)).strftime("%Y-%m")
    next_month = (month_end + timedelta(days=1)).strftime("%Y-%m")
    return month_start, month_end, month_start.strftime("%Y-%m"), prev_month, next_month


def _lang(request: Request) -> str:
    """İstek dili. Cookie > DB setting > default."""
    lang = request.cookies.get("lang")
    if not lang:
        lang = db.get_setting("language", DEFAULT_LANG)
    return lang if lang in SUPPORTED_LANGUAGES else DEFAULT_LANG


def _ctx(request: Request, **extra) -> dict:
    """Template context oluştur."""
    lang = _lang(request)
    tr = load_translations(lang)
    ctx = {
        "request": request,
        "lang": lang,
        "langs": SUPPORTED_LANGUAGES,
        "t": tr,
        "alert_count": db.get_unacknowledged_count(),
        **extra,
    }
    # lazy evaluate cloudflared URL; if syslog not available this is None
    ctx["cloudflare_url"] = get_cloudflared_url()
    return ctx


def _is_htmx(request: Request) -> bool:
    return request.headers.get("HX-Request") == "true"


def _launch_update_command(*args: str) -> None:
    script_path = ROOT_DIR / "update_and_restart.sh"
    log_path = ROOT_DIR / "logs" / "update-command.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    log_handle = open(log_path, "a", encoding="utf-8")
    try:
        subprocess.Popen(
            ["/bin/bash", str(script_path), *args],
            cwd=str(ROOT_DIR),
            stdout=log_handle,
            stderr=log_handle,
            start_new_session=True,
            close_fds=True,
        )
    finally:
        log_handle.close()


# ============================================================ Pages
@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("base.html", {
        **_ctx(request),
        "page": "dashboard",
        "page_content": "partials/dashboard.html",
    })


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page(request: Request):
    ctx = _ctx(request, page="dashboard")
    if _is_htmx(request):
        return templates.TemplateResponse("partials/dashboard.html", ctx)
    ctx["page_content"] = "partials/dashboard.html"
    return templates.TemplateResponse("base.html", ctx)


@app.get("/records", response_class=HTMLResponse)
async def records_page(request: Request,
                       year: Optional[str] = Query(None),
                       month: Optional[str] = Query(None),
                       day: Optional[str] = Query(None)):
    view = "years"
    years = []
    months = []
    summaries = []
    year_label = None
    month_label = None
    day_label = None
    
    # Quick stats (always show at top level)
    quick_stats = {
        "today": db.get_today_count(),
        "week": db.get_week_count(),
        "month": db.get_month_count(),
        "all_time": db.get_all_time_count(),
    }
    
    # Stats section data
    stats_data = {}

    if day:
        # Day view: show daily details + hourly stats
        day_stats = db.get_day_stats(day)
        day_label = day
        month_label = day[:-3]
        year_label = day[:4]
        summaries = [day_stats["daily"]] if day_stats["daily"] else []
        stats_data = {
            "hourly_dist": json.dumps(day_stats["hourly_dist"]),
        }
        view = "day"
    elif month:
        # Month view: show daily summaries for the month + daily trend
        month_start, month_end, month_key, _, _ = _month_bounds(month)
        month_label = month_key
        year_label = str(month_start.year)
        summaries = db.get_daily_summaries(
            start_date=month_start.isoformat(),
            end_date=month_end.isoformat(),
            limit=1000,
        )
        month_stats = db.get_month_stats(month)
        stats_data = {
            "month_total": month_stats["total"],
            "month_days": month_stats["days_with_counts"],
            "month_avg": month_stats["avg_count"],
            "month_peak": month_stats["peak"],
            "daily_trend": json.dumps(month_stats["daily_trend"]),
        }
        view = "month"
    elif year:
        # Year view: show months + monthly trend for the year
        months = db.get_monthly_summaries(year)
        year_label = year
        year_stats = db.get_year_stats(year)
        # Get monthly trend for year
        monthly_rows = db.conn.execute("""
            SELECT strftime('%Y-%m',date) as month, SUM(total_count) as count
            FROM daily_summaries WHERE date LIKE ?
            GROUP BY month ORDER BY month
        """, (f"{year}%",)).fetchall()
        monthly_trend = json.dumps([{"month": r[0], "count": r[1]} for r in monthly_rows])
        stats_data = {
            "year_total": year_stats["total"],
            "monthly_trend": monthly_trend,
        }
        view = "year"
    else:
        # Top view: show years + yearly stats
        years = db.get_yearly_summaries()
        # Get yearly trend
        yearly_rows = db.conn.execute("""
            SELECT substr(date,1,4) as year, SUM(total_count) as count
            FROM daily_summaries GROUP BY year ORDER BY year
        """).fetchall()
        yearly_trend = json.dumps([{"year": r[0], "count": r[1]} for r in yearly_rows])
        stats_data = {
            "yearly_trend": yearly_trend,
        }
        view = "years"

    ctx = _ctx(request,
               page="records",
               view=view,
               years=years,
               months=months,
               summaries=summaries,
               year_label=year_label,
               month_label=month_label,
               day_label=day_label,
               quick_stats=quick_stats,
               **stats_data)
    if _is_htmx(request):
        return templates.TemplateResponse("partials/records.html", ctx)
    ctx["page_content"] = "partials/records.html"
    return templates.TemplateResponse("base.html", ctx)


@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    # List available models from disk
    models_dir = ROOT_DIR / "models"
    available_models = []
    if models_dir.exists():
        for f in sorted(models_dir.rglob("*.pt")):
            available_models.append({
                "path": str(f.relative_to(ROOT_DIR)),
                "name": f.name,
                "type": "pytorch",
            })
        for f in sorted(models_dir.rglob("best.xml")):
            d = f.parent
            available_models.append({
                "path": str(d.relative_to(ROOT_DIR)),
                "name": d.name,
                "type": "openvino",
            })

    ctx = _ctx(
        request,
        page="settings",
        settings_all=db.get_all_settings_detailed(),
        settings=db.get_settings(),
        versions=db.get_versions(),
        active_version=db.get_active_version(),
        available_models=available_models,
        current_version=display_version(read_version()),
    )
    if _is_htmx(request):
        return templates.TemplateResponse("partials/settings.html", ctx)
    ctx["page_content"] = "partials/settings.html"
    return templates.TemplateResponse("base.html", ctx)


@app.get("/logs", response_class=HTMLResponse)
async def logs_page(request: Request):
    date_str = request.query_params.get("date", date.today().isoformat())
    events = db.get_events(date_str=date_str, limit=200)
    alerts = db.get_alerts(limit=50)
    ctx = _ctx(
        request,
        page="logs",
        events=events,
        alerts=alerts,
        log_date=date_str,
    )
    if _is_htmx(request):
        return templates.TemplateResponse("partials/logs.html", ctx)
    ctx["page_content"] = "partials/logs.html"
    return templates.TemplateResponse("base.html", ctx)


# ============================================================ API: Pipeline
@app.post("/api/pipeline/start")
async def api_start(request: Request):
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    source = body.get("source")
    result = pipeline.start(source=source)
    _record_schedule_state("manual_start", "manual_start", result.get("ok", False), result.get("error", ""))
    return JSONResponse(result)


@app.post("/api/pipeline/stop")
async def api_stop():
    result = pipeline.stop()
    _record_schedule_state("manual_stop", "manual_stop", result.get("ok", False), result.get("error", ""))
    return JSONResponse(result)


@app.post("/api/pipeline/pause")
async def api_pause():
    return JSONResponse(pipeline.pause())


@app.post("/api/pipeline/resume")
async def api_resume():
    return JSONResponse(pipeline.resume())


@app.post("/api/pipeline/reset")
async def api_reset():
    return JSONResponse(pipeline.reset_count())


@app.post("/api/pipeline/debug")
async def api_debug():
    v = pipeline.toggle_debug()
    return JSONResponse({"ok": True, "debug": v})


@app.get("/api/pipeline/status")
async def api_status():
    return JSONResponse(_status_payload())


@app.get("/api/pipeline/events")
async def api_events():
    return JSONResponse(pipeline.get_recent_events())


# ============================================================ API: Stream
@app.get("/api/stream")
async def video_stream():
    return StreamingResponse(
        pipeline.frame_generator(),
        media_type="multipart/x-mixed-replace; boundary=frame",
    )


# ============================================================ API: Records
@app.get("/api/sessions")
async def api_sessions(limit: int = 50, offset: int = 0):
    return JSONResponse(db.get_sessions(limit, offset))


@app.delete("/api/sessions/{session_id}")
async def api_delete_session(session_id: int):
    db.delete_session(session_id)
    return JSONResponse({"ok": True})


@app.get("/api/events")
async def api_get_events(session_id: int = None,
                         date_str: str = None,
                         limit: int = 100):
    return JSONResponse(
        db.get_events(session_id=session_id, date_str=date_str, limit=limit)
    )


@app.get("/api/daily")
async def api_daily(start: str = None, end: str = None):
    return JSONResponse(db.get_daily_summaries(start, end))


@app.delete("/api/daily/{date_str}")
async def api_delete_daily(date_str: str):
    db.delete_daily(date_str)
    return JSONResponse({"ok": True})


@app.post("/api/daily/{date_str}/reset")
async def api_reset_daily(date_str: str):
    db.reset_daily(date_str)
    return JSONResponse({"ok": True})


# ============================================================ API: Stats
@app.get("/api/stats")
async def api_stats(days: int = 30):
    return JSONResponse({
        "today": db.get_today_count(),
        "week": db.get_week_count(),
        "month": db.get_month_count(),
        "all_time": db.get_all_time_count(),
        "daily_avg": db.get_daily_average(days),
        "peak_day": db.get_peak_day(),
        "daily_trend": db.get_daily_trend(days),
        "monthly_trend": db.get_monthly_trend(),
        "hourly_dist": db.get_hourly_distribution(),
    })


@app.get("/api/stats/today")
async def api_stats_today():
    return JSONResponse({
        "count": db.get_today_count(),
    })


# ============================================================ API: Settings
@app.get("/api/settings")
async def api_get_settings(category: str = None):
    return JSONResponse(db.get_settings(category))


@app.post("/api/settings")
async def api_save_settings(request: Request):
    body = await request.json()
    if "camera_active_start" in body:
        body["camera_active_start"] = PipelineManager.normalize_schedule_value(
            body.get("camera_active_start"),
            PipelineManager.DEFAULT_CAMERA_ACTIVE_START,
        )
    if "camera_active_end" in body:
        body["camera_active_end"] = PipelineManager.normalize_schedule_value(
            body.get("camera_active_end"),
            PipelineManager.DEFAULT_CAMERA_ACTIVE_END,
        )
    db.set_settings_bulk(body)
    schedule_result = None
    if "camera_active_start" in body or "camera_active_end" in body:
        schedule_result = _enforce_camera_schedule(reason="settings_updated")
    return JSONResponse({"ok": True, "schedule_result": schedule_result})


@app.post("/api/settings/language")
async def api_set_language(request: Request):
    body = await request.json()
    lang = body.get("language", "tr")
    db.set_setting("language", lang)
    response = JSONResponse({"ok": True, "language": lang})
    response.set_cookie("lang", lang, max_age=365 * 24 * 3600)
    return response


# ============================================================ API: Alerts
@app.get("/api/alerts")
async def api_alerts(unack_only: bool = False):
    return JSONResponse(db.get_alerts(unack_only))


@app.post("/api/alerts/{alert_id}/ack")
async def api_ack_alert(alert_id: int):
    db.acknowledge_alert(alert_id)
    return JSONResponse({"ok": True})


@app.post("/api/alerts/ack-all")
async def api_ack_all_alerts():
    db.acknowledge_all_alerts()
    return JSONResponse({"ok": True})


@app.get("/api/alerts/count")
async def api_alert_count():
    return JSONResponse({"count": db.get_unacknowledged_count()})


# ============================================================ API: Export
@app.get("/api/export/csv")
async def export_csv(date_str: str = None):
    events = db.get_events(date_str=date_str, limit=100000)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "id", "session_id", "timestamp", "track_id",
        "cx", "cy", "x1", "y1", "x2", "y2",
        "confidence", "running_total",
    ])
    for e in events:
        writer.writerow([
            e["id"], e["session_id"], e["timestamp"],
            e["track_id"], e["cx"], e["cy"],
            e["x1"], e["y1"], e["x2"], e["y2"],
            e["confidence"], e["running_total"],
        ])

    fname = f"yumurta_rapor_{date_str or 'tum'}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.get("/api/export/excel")
async def export_excel(date_str: str = None):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "openpyxl yüklü değil")

    events = db.get_events(date_str=date_str, limit=100000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sayım Olayları"
    headers = [
        "ID", "Oturum", "Zaman", "Track ID",
        "CX", "CY", "X1", "Y1", "X2", "Y2",
        "Güven", "Toplam",
    ]
    ws.append(headers)
    for e in events:
        ws.append([
            e["id"], e["session_id"], e["timestamp"],
            e["track_id"], e["cx"], e["cy"],
            e["x1"], e["y1"], e["x2"], e["y2"],
            e["confidence"], e["running_total"],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"yumurta_rapor_{date_str or 'tum'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.get("/api/export/pdf")
async def export_pdf(date_str: str = None):
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(500, "fpdf2 yüklü değil")

    events = db.get_events(date_str=date_str, limit=10000)
    summary = db.get_daily_summary(date_str) if date_str else None

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Yumurta Sayim Raporu", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, f"Tarih: {date_str or 'Tum zamanlar'}", ln=True)
    if summary:
        pdf.cell(0, 8,
                 f"Toplam: {summary['total_count']}  |  "
                 f"Oturum: {summary['session_count']}", ln=True)
    pdf.ln(5)

    # Table header
    pdf.set_font("Helvetica", "B", 8)
    col_w = [15, 35, 20, 20, 20, 30]
    headers = ["#", "Zaman", "Track", "CX,CY", "Guven", "Toplam"]
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for idx, e in enumerate(events[:500]):
        pdf.cell(col_w[0], 6, str(idx + 1), border=1)
        pdf.cell(col_w[1], 6, str(e["timestamp"])[11:], border=1)
        pdf.cell(col_w[2], 6, str(e["track_id"] or ""), border=1)
        pdf.cell(col_w[3], 6, f"{e['cx']},{e['cy']}", border=1)
        pdf.cell(col_w[4], 6, f"{e['confidence']:.2f}", border=1)
        pdf.cell(col_w[5], 6, str(e["running_total"]), border=1)
        pdf.ln()

    buf = io.BytesIO(pdf.output())
    fname = f"yumurta_rapor_{date_str or 'tum'}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ============================================================ API: Import
@app.post("/api/import/csv")
async def import_csv(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    events = []
    for row in reader:
        events.append({
            "timestamp": row.get("timestamp", ""),
            "track_id": int(row.get("track_id", 0) or 0),
            "cx": int(row.get("cx", 0) or 0),
            "cy": int(row.get("cy", 0) or 0),
            "x1": int(row.get("x1", 0) or 0),
            "y1": int(row.get("y1", 0) or 0),
            "x2": int(row.get("x2", 0) or 0),
            "y2": int(row.get("y2", 0) or 0),
            "confidence": float(row.get("confidence", 0) or 0),
            "running_total": int(row.get("running_total", 0) or 0),
        })
    count = db.import_count_events(events)
    return JSONResponse({"ok": True, "imported": count})


# ============================================================ WebSocket
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    ws_connections.append(websocket)
    try:
        while True:
            # Send status + events
            status = pipeline.get_status()
            new_events = pipeline.get_new_events(max_count=10)
            msg = {
                "status": _status_payload(),
                "events": new_events,
                "alert_count": db.get_unacknowledged_count(),
            }
            await websocket.send_json(msg)
            await asyncio.sleep(0.5)
    except WebSocketDisconnect:
        pass
    except Exception:
        pass
    finally:
        if websocket in ws_connections:
            ws_connections.remove(websocket)


# ============================================================ Version/Update
@app.get("/api/versions")
async def api_versions():
    return JSONResponse(db.get_versions())


@app.post("/api/versions")
async def api_add_version(request: Request):
    body = await request.json()
    db.add_version(
        body["version"],
        body.get("changelog"),
        body.get("backup_path"),
    )
    return JSONResponse({"ok": True})


@app.post("/api/versions/{vid}/rollback")
async def api_rollback(vid: int):
    ok = db.rollback_version(vid)
    return JSONResponse({"ok": ok})


@app.get("/api/update/status")
async def api_update_status():
    return JSONResponse(update_manager.get_status())


@app.get("/api/update/releases")
async def api_update_releases(include_prerelease: Optional[bool] = None):
    try:
        releases = update_manager.list_releases(include_prerelease=include_prerelease)
        return JSONResponse({"ok": True, "releases": releases})
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc), "releases": []}, status_code=503)


@app.post("/api/update/check")
async def api_update_check(request: Request):
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    result = update_manager.check_for_updates(
        notify=body.get("notify", True),
        include_prerelease=body.get("include_prerelease"),
    )
    return JSONResponse({"ok": result.get("error") is None, **result})


@app.post("/api/update/install")
async def api_update_install(request: Request):
    body = await request.json()
    status = update_manager.get_status()
    if status.get("busy"):
        return JSONResponse({"ok": False, "error": "Başka bir güncelleme işlemi çalışıyor"}, status_code=409)
    args = ["install"]
    version = body.get("version")
    if version:
        args.extend(["--version", str(version)])
    if body.get("restart_after", True):
        args.append("--restart")
    _launch_update_command(*args)
    return JSONResponse({"ok": True, "started": True, "target_version": version})


@app.post("/api/update/rollback")
async def api_update_rollback(request: Request):
    body = await request.json()
    version = body.get("version")
    if not version:
        return JSONResponse({"ok": False, "error": "Sürüm bilgisi gerekli"}, status_code=400)
    status = update_manager.get_status()
    if status.get("busy"):
        return JSONResponse({"ok": False, "error": "Başka bir güncelleme işlemi çalışıyor"}, status_code=409)
    args = ["rollback", "--version", str(version)]
    if body.get("restart_after", True):
        args.append("--restart")
    _launch_update_command(*args)
    return JSONResponse({"ok": True, "started": True, "target_version": version})


@app.post("/api/update/restart")
async def api_update_restart():
    status = update_manager.get_status()
    if status.get("busy"):
        return JSONResponse({"ok": False, "error": "Önce aktif güncelleme işlemi bitsin"}, status_code=409)
    _launch_update_command("restart")
    return JSONResponse({"ok": True, "started": True})


# ============================================================ API: Models
@app.get("/api/models")
async def api_list_models():
    """Mevcut model dosyalarını listele."""
    models_dir = ROOT_DIR / "models"
    result = []
    if models_dir.exists():
        # PyTorch models (.pt)
        for f in sorted(models_dir.rglob("*.pt")):
            try:
                size_mb = round(f.stat().st_size / 1024 / 1024, 1)
            except Exception:
                size_mb = 0
            result.append({
                "path": str(f.relative_to(ROOT_DIR)),
                "name": f.name,
                "type": "pytorch",
                "size_mb": size_mb,
            })
        # OpenVINO model directories (contain best.xml)
        for f in sorted(models_dir.rglob("best.xml")):
            d = f.parent
            try:
                size_mb = round(
                    sum(x.stat().st_size for x in d.iterdir() if x.is_file())
                    / 1024 / 1024, 1
                )
            except Exception:
                size_mb = 0
            result.append({
                "path": str(d.relative_to(ROOT_DIR)),
                "name": d.name,
                "type": "openvino",
                "size_mb": size_mb,
            })
    return JSONResponse(result)


# ============================================================ API: Camera Sources
@app.get("/api/sources/cameras")
async def api_list_cameras():
    """Kullanılabilir kamera cihazlarını listele."""
    import cv2 as _cv2
    cameras = []
    for i in range(6):
        cap = _cv2.VideoCapture(i)
        if cap.isOpened():
            w = int(cap.get(_cv2.CAP_PROP_FRAME_WIDTH))
            h = int(cap.get(_cv2.CAP_PROP_FRAME_HEIGHT))
            cameras.append({
                "index": i,
                "value": str(i),
                "label": f"Kamera {i} ({w}×{h})",
            })
            cap.release()
    return JSONResponse(cameras)


# ============================================================ API: Optimize
import threading as _threading

_optimize_state: Dict = {
    "running": False, "progress": 0, "total": 0,
    "current": "", "results": [], "error": None, "done": False,
    "last_accuracy": 0, "last_fps": 0, "best_accuracy": 0, "best_fps": 0
}
_optimize_lock = _threading.Lock()


@app.get("/api/optimize/status")
async def api_optimize_status():
    with _optimize_lock:
        # JSONResponse'a geçmeden önce yüzeysel kopya al (race condition önlemi)
        return JSONResponse(dict(_optimize_state))


@app.post("/api/optimize/start")
async def api_optimize_start(request: Request):
    global _optimize_state
    with _optimize_lock:
        if _optimize_state.get("running"):
            return JSONResponse({"ok": False, "error": "Optimizasyon zaten çalışıyor"})
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    video_path = body.get("video", str(ROOT_DIR / "data" / "demo.mp4"))
    expected = int(body.get("expected", 200))
    with _optimize_lock:
        _optimize_state = {
            "running": True, "progress": 0, "total": 0,
            "current": "Başlatılıyor...", "results": [],
            "error": None, "done": False,
            "last_accuracy": 0, "last_fps": 0, "best_accuracy": 0, "best_fps": 0
        }
    asyncio.create_task(_optimize_background(video_path, expected))
    return JSONResponse({"ok": True})


@app.post("/api/optimize/cancel")
async def api_optimize_cancel():
    global _optimize_state
    with _optimize_lock:
        _optimize_state["running"] = False
    return JSONResponse({"ok": True})


@app.post("/api/optimize/apply")
async def api_optimize_apply(request: Request):
    # gelen tüm anahtarları ayarlara yaz
    body = await request.json()
    settings_to_save = {k: str(v) for k, v in body.items()}
    db.set_settings_bulk(settings_to_save)
    return JSONResponse({"ok": True, "applied": settings_to_save})


async def _optimize_background(video_path: str, expected: int):
    """Bayesian Optimization (Optuna) kullanarak parametreleri optimize et."""
    global _optimize_state
    import optuna

    # Optuna loglarını sustur (sadece hata mesajları)
    optuna.logging.set_verbosity(optuna.logging.WARNING)

    # Toplam deneme sayısı (Bütçe)
    n_trials = 100
    with _optimize_lock:
        _optimize_state["total"] = n_trials
        _optimize_state["results"] = []  # Listeyi temizle

    loop = asyncio.get_event_loop()
    results = []

    def objective(trial):
        # Durdurma kontrolü (Kullanıcı iptal ettiyse)
        with _optimize_lock:
            if not _optimize_state.get("running"):
                trial.study.stop()
                return 0.0

        # Parametre Uzayı Tanımı (Bayesian Arama için)
        params = {
            "conf_threshold": trial.suggest_float("conf_threshold", 0.15, 0.50, step=0.05),
            "iou_threshold": trial.suggest_float("iou_threshold", 0.20, 0.60, step=0.05),
            "line_position": trial.suggest_float("line_position", 0.30, 0.70, step=0.05),
            "tracker_type": trial.suggest_categorical("tracker_type", ["bytetrack", "botsort"]),
            "imgsz": trial.suggest_categorical("imgsz", [320, 480, 640]),
            "track_buffer": trial.suggest_int("track_buffer", 30, 150, step=30),
            "match_thresh": trial.suggest_float("match_thresh", 0.70, 0.95, step=0.05),
            "skip_frames": trial.suggest_int("skip_frames", 0, 2),
            "crop_ud": trial.suggest_int("crop_ud", 0, 20, step=5),
            "crop_lr": trial.suggest_int("crop_lr", 0, 20, step=5),
            "enable_clahe": trial.suggest_categorical("enable_clahe", [True, False]),
            "enable_stabilization": trial.suggest_categorical("enable_stabilization", [True, False]),
            "roi_top": trial.suggest_float("roi_top", 0.15, 0.40, step=0.05),
            "roi_bottom": trial.suggest_float("roi_bottom", 0.60, 0.85, step=0.05),
        }

        # UI Güncelleme (Hangi kombinasyonun denendiğini göster)
        with _optimize_lock:
            _optimize_state["progress"] = trial.number
            desc = (
                f"Deneme {trial.number+1}/{n_trials}: "
                f"conf={params['conf_threshold']} imgsz={params['imgsz']} tracker={params['tracker_type']}"
            )
            _optimize_state["current"] = desc

        try:
            # Senkron fonksiyonu çalıştır
            result = _count_video_sync(video_path, params)
            count = result.get("count", -1)
            
            if count >= 0:
                error = abs(count - expected)
                accuracy = round(max(0, 100 - (error / max(1, expected)) * 100), 1)
                fps = result.get("fps", 0)

                out = {
                    "params": params.copy(),
                    "count": count,
                    "expected": expected,
                    "error": error,
                    "accuracy": accuracy,
                    "fps": fps,
                    "conf": params["conf_threshold"],
                    "iou": params["iou_threshold"],
                    "line_position": params["line_position"],
                    "tracker_type": params["tracker_type"]
                }
                results.append(out)

                # Sıralama ve UI Güncelleme
                with _optimize_lock:
                    _optimize_state["last_accuracy"] = accuracy
                    _optimize_state["last_fps"] = fps
                    # Tüm sonuçları tut (15 sınırı kaldırıldı)
                    results.sort(key=lambda x: (-x["accuracy"], -x["fps"]))
                    for rank_i, r in enumerate(results):
                        r["rank"] = rank_i + 1
                    _optimize_state["results"] = results.copy() 
                    _optimize_state["best_accuracy"] = results[0]["accuracy"] if results else 0
                    _optimize_state["best_fps"] = results[0]["fps"] if results else 0

                # Bayesian Hedef: Doğruluğu maksimize et, yüksek FPS'i ödüllendir
                # score = -accuracy + weight * -fps (minimize modunda)
                score = -accuracy - (0.01 * fps)
                return score
            
            return 1000.0  # Hatalı deneme için kötü skor
        except Exception:
            return 1000.0

    try:
        # Optuna Study oluştur (TPESampler varsayılan Bayesian örneklendiricidir)
        study = optuna.create_study(direction="minimize")
        # Optuna'yı ayrı bir thread'de çalıştır (IO/CPU bound karmaşasını önlemek adına executor kullanılır)
        await loop.run_in_executor(None, study.optimize, objective, n_trials)
    except Exception as e:
        with _optimize_lock:
            _optimize_state["error"] = str(e)

    with _optimize_lock:
        _is_cancelled = not _optimize_state.get("running")
        _optimize_state["running"] = False
        _optimize_state["done"] = True
        _optimize_state["progress"] = n_trials
        _optimize_state["current"] = "İptal Edildi" if _is_cancelled else "Tamamlandı"


def _count_video_sync(
    video_path: str, params: dict,
) -> dict:
    """Videoyu minimal pipeline ile işle ve toplam sayımı döndür (sync, threadsafe).

    `params` sözlüğü optimize edilmiş değerleri içerir; fonksiyon bu anahtarları
    uygun konfigürasyon objelerine geçirir ve pipeline'ı çalıştırır.
    """
    import cv2 as _cv2
    import time as _time

    try:
        from egg_counter.config import (
            DetectorConfig, TrackerConfig, CounterConfig, PipelineConfig,
        )
        from egg_counter.detector import EggDetector
        from egg_counter.tracker import TrackManager
        from egg_counter.counter import CountingLine

        s = db.get_settings()
        model_path = s.get("model_path", "models/yolo26n_mod/a3_best_openvino_model")

        # detector
        cfg_det = DetectorConfig(
            model_path=model_path,
            imgsz=params.get("imgsz", 480),
            conf_threshold=params.get("conf_threshold", 0.30),
            iou_threshold=params.get("iou_threshold", 0.45),
            enable_clahe=params.get("enable_clahe", False),
            enable_stabilization=params.get("enable_stabilization", False),
        )

        # tracker
        cfg_trk = TrackerConfig(
            tracker_type=params.get("tracker_type", "bytetrack"),
            track_buffer=params.get("track_buffer", 90),
            match_thresh=params.get("match_thresh", 0.85),
        )

        # counter
        cfg_ctr = CounterConfig(
            line_position=params.get("line_position", 0.5),
            direction="both",
            roi_top_position=params.get("roi_top", 0.35),
            roi_bottom_position=params.get("roi_bottom", 0.65),
        )

        # pipeline
        cfg_pipe = PipelineConfig(
            skip_frames=params.get("skip_frames", 0),
            crop_ud=params.get("crop_ud", 0),
            crop_lr=params.get("crop_lr", 0),
        )

        # video mı yoksa resim mi olduğuna bak
        path = str(video_path)
        ext = path.lower().rsplit('.', 1)[-1] if '.' in path else ''
        is_image = ext in ('jpg', 'jpeg', 'png', 'bmp', 'tif', 'tiff')

        if is_image:
            t0 = _time.time()
            frame = _cv2.imread(path)
            if frame is None:
                return {"count": -1, "error": f"Resim açılamadı: {path}"}

            detector = EggDetector(cfg_det, cfg_trk)

            # tespitleri al ve say
            raw = detector.detect_and_track(frame)
            detections = detector.parse_results(raw)
            total = len(detections)

            elapsed = max(0.001, _time.time() - t0)
            return {"count": total, "fps": round(1.0 / elapsed, 1), "frames": 1}
        else:
            cap = _cv2.VideoCapture(path)
            if not cap.isOpened():
                return {"count": -1, "error": f"Video açılamadı: {path}"}

            fps_v = cap.get(_cv2.CAP_PROP_FPS) or 30
            h = int(cap.get(_cv2.CAP_PROP_FRAME_HEIGHT))

            detector = EggDetector(cfg_det, cfg_trk)
            track_mgr = TrackManager(cfg_trk, cfg_ctr, trail_length=20)
            counter = CountingLine(cfg_ctr, frame_height=h)

            total = 0
            frames = 0
            t0 = _time.time()

            while True:
                ret, frame = cap.read()
                if not ret:
                    break
                raw = detector.detect_and_track(frame)
                detections = detector.parse_results(raw)
                enriched = track_mgr.update(detections)
                events = counter.check_crossings(enriched, track_mgr)
                total += len(events)
                frames += 1

            cap.release()
            elapsed = max(0.01, _time.time() - t0)
            return {"count": total, "fps": round(frames / elapsed, 1), "frames": frames}

    except Exception as e:
        return {"count": -1, "error": str(e)}
